import os
import sys
import numpy as np
import openpyxl
from openpyxl import Workbook
import networkx as nx
import scipy.io as scio
from scipy.sparse import triu
from __CONFIGURATION import net_names
from __UTIL import load_edge_list, save_edge_list
from oriNet import highSchool, primarySchool, facebook100, SyntheticSSBM


class EmpiricalNet:
    def __init__(self, net_name, force_calc=False):
        """
        Prepare Empirical Network for rapid read and use.
        :param net_name: network name
        """
        self.net_name = net_name
        self.force_calc = force_calc
        if self.net_name.startswith('facebook100'):
            self.path_name = 'facebook100'
        else:
            self.path_name = self.net_name
        self.net_path = "./net_data/" + self.path_name + "/"
        self.ori_path = self.net_path + "origin/"  # TODO NEED PRE SET Origin Path
        self.mat_path = self.net_path + "mat/"
        self.adj_path = self.net_path + "adj/"
        self.gexf_path = self.net_path + "gexf/"
        self.info_path = self.net_path + self.net_name + "_net_info.txt"
        self._make_path()
        self._make_oriNet()
        # set the variable
        self.is_sparse = False  # TODO sparse network(Future in Subclass), from npz to networkX too slow.
        self.mat = None  # run get_mat() to get adjacent matrix (CSR Sparse Matrix)
        self.n = 0  # run get_mat() to get node number
        self.edges = None  # run get_adj() to get adjacent table
        # run get_node_attr() to get node attributes: {"attrName": [attrV[0], ...], ...}
        # and node attributes meta {"attrName": {attrV:attrMeta, ...}, ...}
        self.nodeAttr = None
        self.nodeAttrMeta = None
        self.nxGraph = None  # run get_nx() to get networkX Graph
        self._prepare()

    def _make_path(self):
        # make directories if it doesn't exist
        os.makedirs(self.mat_path) if os.path.exists(self.mat_path) is False else None
        os.makedirs(self.adj_path) if os.path.exists(self.adj_path) is False else None
        os.makedirs(self.gexf_path) if os.path.exists(self.gexf_path) is False else None

    def _make_oriNet(self):
        if self.net_name == 'highSchool':
            self.ori_net = highSchool(self.ori_path)
        elif self.net_name == 'primarySchool':
            self.ori_net = primarySchool(self.ori_path)
        elif self.net_name.startswith('facebook100'):
            region = self.net_name.split('_')[1]
            self.ori_net = facebook100(self.ori_path, region)
        else:
            self.ori_net = None
            print(self.net_name, "is wrong!")
            sys.exit()

    def _prepare(self):
        self.get_mat()
        self.get_edges()
        self.get_node_attr()
        self.get_nx()
        self.show_net_info()

    def get_mat(self):
        """
        get adjacent matrix
        :return: CSR Sparse Matrix
        """
        if self.mat is not None:
            print(self.net_name + " adjacent matrix already loaded!")
            return self.mat
        mat_file_path = self.mat_path + self.net_name + ".mat"
        if os.path.exists(mat_file_path) and self.force_calc is False:
            print(self.net_name + " adjacent matrix already prepared!")
            mat = scio.loadmat(mat_file_path)['net']
        else:
            mat = self.ori_net.get_mat()
            scio.savemat(mat_file_path, {'net': mat})
        self.mat = mat
        self.n = np.shape(self.mat)[0]
        return self.mat

    def get_node_attr(self):
        if self.nodeAttr is not None and self.nodeAttrMeta is not None:
            print(self.net_name + " node attribute already loaded!")
            return self.nodeAttr, self.nodeAttrMeta
        self.nodeAttr, self.nodeAttrMeta = self.ori_net.get_node_attr()
        return self.nodeAttr, self.nodeAttrMeta

    def get_edges(self):
        """
        get adjacent table (from up triangle)
        :return: [(node1,node2), (node1,node3),...]
        """
        if self.edges is not None:
            print(self.net_name + " edges already loaded!")
            return self.edges
        adj_file_path = self.adj_path + self.net_name + ".txt"
        if os.path.exists(adj_file_path) and self.force_calc is False:
            print(self.net_name + " edges already prepared!")
            edges = load_edge_list(adj_file_path)
        else:
            mat = self.get_mat()
            edge_coord = triu(mat)  # column first
            x = edge_coord.row
            y = edge_coord.col
            edges = []
            for j in range(len(x)):
                t = (x[j], y[j])
                edges.append(t)
            save_edge_list(edges, adj_file_path)
        self.edges = edges
        return self.edges

    def get_nx(self):
        """
        get networkX graph.
        save and load by gephi file format(gexf)
        :return:
        """
        gexf_file_path = self.gexf_path + self.net_name + ".gexf"
        if os.path.exists(gexf_file_path) and self.force_calc is False:
            g = nx.read_gexf(gexf_file_path)
            print(self.net_name + " networkX graph already prepared!")
        else:
            mat = self.get_mat()
            nodeAttr, nodeAttrMeta = self.get_node_attr()
            g = nx.from_scipy_sparse_matrix(mat)
            g.name = self.net_name
            for n in g.nodes:
                for attrName in nodeAttr.keys():
                    g.nodes[n][attrName] = nodeAttrMeta[attrName][nodeAttr[attrName][n]]
            nx.write_gexf(g, gexf_file_path)
        self.nxGraph = g
        return self.nxGraph

    def show_net_info(self):
        info = ""
        with open(self.info_path, 'w') as info_f:
            # name, nodes, edges, average degree
            nx_net = self.nxGraph
            nx_info = nx.info(nx_net)
            info += nx_info + '\n'
            info += "Is connected: " + str(nx.is_connected(nx_net)) + "\n"
            # attribute
            for attrName in self.nodeAttrMeta:
                if len(self.nodeAttrMeta[attrName]) < 10:
                    info += attrName + '(idmap): ' + str(self.nodeAttrMeta[attrName]) + '\n'
                else:
                    info += attrName + '(idmap): length is ' + str(len(self.nodeAttrMeta[attrName])) + '\n'
            # output to console and file
            print(info)
            info_f.write(info)

    def summary_net_data(self):
        summary_filename = "./net_data/summary.xlsx"
        title = "summary"
        if os.path.exists(summary_filename):
            wb = openpyxl.load_workbook(summary_filename)
            ws = wb[title]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = title
            ws.cell(row=1, column=1, value='name')
            ws.cell(row=1, column=2, value='nodes')
            ws.cell(row=1, column=3, value='edges')
            ws.cell(row=1, column=4, value='average_degree')
            ws.cell(row=1, column=5, value='is_connected')
            ws.cell(row=1, column=6, value='MCC_nodes')
        row_i = 2
        while True:
            c = ws['A' + str(row_i)]
            if c.value is None:
                ws.cell(row=row_i, column=1, value=self.net_name)
                ws.cell(row=row_i, column=2, value=self.nxGraph.number_of_nodes())
                ws.cell(row=row_i, column=3, value=self.nxGraph.number_of_edges())
                ws.cell(row=row_i, column=4, value=sum(dict(self.nxGraph.degree()).values()) / float(
                    self.nxGraph.number_of_nodes()))
                ws.cell(row=row_i, column=5, value=nx.is_connected(self.nxGraph))
                ws.cell(row=row_i, column=6, value=max([len(c) for c in nx.connected_components(self.nxGraph)]))
                break
            elif c.value == self.net_name:
                break
            else:
                pass
            row_i += 1
        wb.save(summary_filename)


class SyntheticNet:
    def __init__(self, net_name, verbose=True):
        self.net_name = net_name
        self.n = 0
        self.mat = None  # run get_mat() to get adjacent matrix (CSR Sparse Matrix)
        self.edges = None  # run get_edges() to get adjacent table
        self.verbose = verbose
        # run get_node_attr() to get node attributes: {"attrName": [attrV[0], ...], ...}
        # and node attributes meta {"attrName": {attrV:attrMeta, ...}, ...}
        self.nodeAttr = None
        self.nodeAttrMeta = None
        self.nxGraph = None  # run get_nx() to get networkX Graph
        self._prepare(verbose)

    def _make_oriNet(self):
        if self.net_name.startswith('SSBM'):
            """
                Symmetric SBM network parameter order: n c k epsilon
            """
            parameters = self.net_name.split("_")
            self.n = int(parameters[1])
            c = float(parameters[2])
            k = int(parameters[3])
            epsilon = float(parameters[4])
            self.ori_net = SyntheticSSBM(self.n, c, k, epsilon)
        else:
            self.ori_net = None
            print(self.net_name, "is wrong!")
            sys.exit()

    def get_mat(self):
        if self.mat is not None:
            if self.verbose:
                print(self.net_name + " adjacent matrix already loaded!")
            return self.mat
        self.mat = self.ori_net.get_mat()
        return self.mat

    def get_node_attr(self):
        if self.nodeAttr is not None and self.nodeAttrMeta is not None:
            if self.verbose:
                print(self.net_name + " node attribute already loaded!")
            return self.nodeAttr, self.nodeAttrMeta
        self.nodeAttr, self.nodeAttrMeta = self.ori_net.get_node_attr()
        return self.nodeAttr, self.nodeAttrMeta

    def get_edges(self):
        if self.edges is not None:
            if self.verbose:
                print(self.net_name + " edges already loaded!")
            return self.edges
        mat = self.get_mat()
        edge_coord = triu(mat)  # column first
        x = edge_coord.row
        y = edge_coord.col
        edges = []
        for j in range(len(x)):
            t = (x[j], y[j])
            edges.append(t)
        self.edges = edges
        return self.edges

    def get_nx(self):
        mat = self.get_mat()
        nodeAttr, nodeAttrMeta = self.get_node_attr()
        g = nx.from_scipy_sparse_matrix(mat)
        g.name = self.net_name
        for n in g.nodes:
            for attrName in nodeAttr.keys():
                g.nodes[n][attrName] = nodeAttrMeta[attrName][nodeAttr[attrName][n]]
        self.nxGraph = g
        return self.nxGraph

    def show_net_info(self):
        info = ""
        # name, nodes, edges, average degree
        nx_net = self.nxGraph
        nx_info = nx.info(nx_net)
        info += nx_info + '\n'
        info += "Is connected: " + str(nx.is_connected(nx_net)) + "\n"
        # attribute
        for attrName in self.nodeAttrMeta:
            if len(self.nodeAttrMeta[attrName]) < 10:
                info += attrName + '(idmap): ' + str(self.nodeAttrMeta[attrName]) + '\n'
            else:
                info += attrName + '(idmap): length is ' + str(len(self.nodeAttrMeta[attrName])) + '\n'
        # output to console
        print(info)

    def _prepare(self, verbose):
        self._make_oriNet()
        self.get_mat()
        self.get_edges()
        self.get_node_attr()
        self.get_nx()
        if verbose:
            self.show_net_info()


if __name__ == '__main__':
    for name in net_names:
        EmpiricalNet(name).summary_net_data()
    # EmpiricalNet("facebook100_FSU53")
    # EmpiricalNet("highSchool")
    # EmpiricalNet("primarySchool")
    # EmpiricalNet("facebook100_American75")
